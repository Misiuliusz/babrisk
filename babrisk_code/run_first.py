# Computing the indicator variables based on whether the original 
# data were written in italic font (uncertain values). 
# See http://www.iisg.nl/hpw/babylon.php for the details.

import xlrd
import openpyxl
import numpy as np
import os

absolute_path = os.path.dirname(__file__)
relative_path_xls = 'OriginalData\\babylonia.xls'
path_xls = os.path.join(absolute_path, relative_path_xls)

wb = xlrd.open_workbook(path_xls, formatting_info=True)
sh = wb.sheet_by_name('Blad1')
nrows = 1499

def blank_test(cell):
    result = (cell=='')|(cell==' ')|(cell=='  ')|(cell=='   ') \
    |(cell=='    ')|(cell=='     ')|(cell=='      ')|(cell=='       ')
    return result

italic_s_b = np.zeros(nrows,)
for row in range(1,nrows):
    font = wb.font_list
    cell_val = sh.cell_value(row,11)
    cell_val_orig = sh.cell_value(row,10)
    cell_xf = wb.xf_list[sh.cell_xf_index(row,11)]
    italic_s_b[row] = int(font[cell_xf.font_index].italic & ~blank_test(cell_val))
    
italic_s_d = np.zeros(nrows,)
for row in range(1,nrows):
    font = wb.font_list
    cell_val = sh.cell_value(row,16)
    cell_val_orig = sh.cell_value(row,15)  
    cell_xf = wb.xf_list[sh.cell_xf_index(row,16)]
    italic_s_d[row] = int(font[cell_xf.font_index].italic & ~blank_test(cell_val))

# This must be .xlsx file (a requirement of xlrd package).
relative_path_xlsx = 'OriginalData\\babylonia.xlsx'
path_xlsx = os.path.join(absolute_path, relative_path_xlsx)
wb = openpyxl.load_workbook(path_xlsx)
ws = wb['Blad1']
nrows = 1499

ws['AN1'] = 'italicBarint'
ws['AO1'] = 'italicDatint'

for row in range(1,nrows):
    row_text = str(row+1)
    ws['AN'+row_text] = italic_s_b[row]
    ws['AO'+row_text] = italic_s_d[row]

wb.save(path_xlsx)